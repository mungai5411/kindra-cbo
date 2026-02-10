"""
Reporting Services
Business logic for generating system reports (PDF/Excel/CSV)
"""

import io
import csv
import logging
from django.utils import timezone
from django.core.files.base import ContentFile
from reportlab.lib.pagesizes import letter, landscape
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

logger = logging.getLogger('kindra_cbo')

class ReportService:
    """
    Service class for generating various system reports
    """

    @staticmethod
    def generate_report_file(report):
        """
        Main entry point to generate the physical file for a report object
        Returns the generated content to avoid re-reading from storage
        """
        try:
            content = None
            if report.format == 'PDF':
                content = ReportService._generate_pdf(report)
            elif report.format == 'EXCEL':
                content = ReportService._generate_excel(report)
            else:
                # Default to CSV if not PDF/Excel
                content = ReportService._generate_csv(report)
            
            report.status = 'COMPLETED'
            report.save()
            logger.info(f"Successfully generated {report.format} for report {report.id}")
            return content
        except Exception as e:
            logger.error(f"Error generating report {report.id}: {str(e)}")
            report.status = 'FAILED'
            report.save()
            raise

    @staticmethod
    def _generate_csv(report):
        from donations.models import Donation
        from volunteers.models import Volunteer
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        if report.report_type == 'DONATION':
            writer.writerow(['Date', 'Donor', 'Amount', 'Currency', 'Method', 'Campaign', 'Status'])
            # Filtering by date range if provided
            qs = Donation.objects.all()
            if report.start_date: qs = qs.filter(donation_date__gte=report.start_date)
            if report.end_date: qs = qs.filter(donation_date__lte=report.end_date)
            
            for d in qs:
                writer.writerow([
                    d.donation_date.strftime('%Y-%m-%d %H:%M'),
                    d.donor_name or str(d.donor or "Anonymous"),
                    d.amount,
                    d.currency,
                    d.payment_method,
                    d.campaign.title if d.campaign else 'General Fund',
                    d.status
                ])
        elif report.report_type == 'VOLUNTEER':
            writer.writerow(['Name', 'Email', 'Phone', 'Role', 'Status', 'Total Hours', 'Join Date'])
            qs = Volunteer.objects.all()
            for v in qs:
                writer.writerow([
                    v.full_name, v.email, v.phone_number, 'VOLUNTEER', 
                    v.status, v.total_hours, v.created_at.strftime('%Y-%m-%d')
                ])
        elif report.report_type == 'CASE':
            from case_management.models import Case
            writer.writerow(['Case Number', 'Title', 'Status', 'Priority', 'Family', 'Assigned To', 'Opened Date'])
            qs = Case.objects.all().select_related('family', 'assigned_to')
            if report.start_date: qs = qs.filter(opened_date__gte=report.start_date)
            if report.end_date: qs = qs.filter(opened_date__lte=report.end_date)
            
            for c in qs:
                writer.writerow([
                    c.case_number,
                    c.title,
                    c.status,
                    c.priority,
                    c.family.family_code,
                    c.assigned_to.get_full_name() if c.assigned_to else 'Unassigned',
                    c.opened_date.strftime('%Y-%m-%d')
                ])
        elif report.report_type == 'SHELTER':
            from shelter_homes.models import ShelterHome
            writer.writerow(['Name', 'Reg Number', 'County', 'Contact', 'Email', 'Phone', 'Capacity', 'Occupancy', 'Status'])
            qs = ShelterHome.objects.all()
            for s in qs:
                writer.writerow([
                    s.name,
                    s.registration_number,
                    s.county,
                    s.contact_person,
                    s.email,
                    s.phone_number,
                    s.total_capacity,
                    s.current_occupancy,
                    s.status
                ])
        
        content = output.getvalue().encode('utf-8')
        return content

    @staticmethod
    def _generate_excel(report):
        from donations.models import Donation
        from volunteers.models import Volunteer

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{report.report_type} Report"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="519755", end_color="519755", fill_type="solid")
        center_align = Alignment(horizontal="center")

        if report.report_type == 'DONATION':
            headers = ['Date', 'Donor', 'Amount', 'Currency', 'Method', 'Campaign', 'Status']
            ws.append(headers)
            
            qs = Donation.objects.all()
            if report.start_date: qs = qs.filter(donation_date__gte=report.start_date)
            if report.end_date: qs = qs.filter(donation_date__lte=report.end_date)

            for d in qs:
                ws.append([
                    d.donation_date.replace(tzinfo=None), # Excel doesn't like tz-aware
                    d.donor_name or str(d.donor or "Anonymous"),
                    float(d.amount),
                    d.currency,
                    d.payment_method,
                    d.campaign.title if d.campaign else 'General Fund',
                    d.status
                ])
        elif report.report_type == 'VOLUNTEER':
            headers = ['Name', 'Email', 'Phone', 'Status', 'Total Hours', 'Join Date']
            ws.append(headers)
            qs = Volunteer.objects.all()
            for v in qs:
                ws.append([
                    v.full_name, v.email, v.phone_number, v.status, 
                    float(v.total_hours), v.created_at.replace(tzinfo=None)
                ])

        # Style headers
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        content = buffer.read()
        return content

    @staticmethod
    def _generate_pdf(report):
        """
        Generate professional PDF report using HTML templates and WeasyPrint
        """
        from django.template.loader import render_to_string
        from weasyprint import HTML
        from django.conf import settings
        from donations.models import Donation
        from volunteers.models import Volunteer
        import os
        
        try:
            # Logo path (optional)
            logo_path = os.path.join(settings.BASE_DIR, 'donations', 'static', 'donations', 'images', 'logo.jpg')
            if not os.path.exists(logo_path):
                logo_path = None
            
            # Template selection based on report type
            template_map = {
                'DONATION': 'reporting/donation_report.html',
                'VOLUNTEER': 'reporting/volunteer_report.html',
                'CASE': 'reporting/case_report.html',
                'FINANCIAL': 'reporting/financial_report.html',
                'COMPLIANCE': 'reporting/compliance_report.html',
                'SHELTER': 'reporting/shelter_report.html',
            }
            
            template_name = template_map.get(report.report_type, 'reporting/base_report.html')
            
            # Prepare data based on report type
            context = {
                'report': report,
                'logo_path': logo_path,
            }
            
            if report.report_type == 'DONATION':
                qs = Donation.objects.all().select_related('donor', 'campaign')
                if report.start_date: qs = qs.filter(donation_date__gte=report.start_date)
                if report.end_date: qs = qs.filter(donation_date__lte=report.end_date)
                
                donations = list(qs[:500])  # Limit for performance
                context['donations'] = donations
                context['summary'] = {
                    'total_donations': len(donations),
                    'total_amount': sum(d.amount for d in donations),
                    'donors_count': len(set(d.donor_id for d in donations if d.donor_id)),
                }
                
            elif report.report_type == 'VOLUNTEER':
                qs = Volunteer.objects.all()
                volunteers = list(qs[:500])
                context['volunteers'] = volunteers
                context['summary'] = {
                    'total_volunteers': len(volunteers),
                    'total_hours': sum(v.total_hours for v in volunteers),
                    'active_count': len([v for v in volunteers if v.status == 'ACTIVE']),
                }

            elif report.report_type == 'SHELTER':
                from shelter_homes.models import ShelterHome
                qs = ShelterHome.objects.all()
                shelters = list(qs)
                context['shelters'] = shelters
                context['summary'] = {
                    'total_shelters': len(shelters),
                    'total_capacity': sum(s.total_capacity for s in shelters),
                    'total_occupancy': sum(s.current_occupancy for s in shelters),
                }

            elif report.report_type == 'CASE':
                from case_management.models import Case
                qs = Case.objects.all().select_related('family', 'assigned_to')
                if report.start_date: qs = qs.filter(opened_date__gte=report.start_date)
                if report.end_date: qs = qs.filter(opened_date__lte=report.end_date)
                
                cases = list(qs[:500])
                context['cases'] = cases
                context['summary'] = {
                    'total_cases': len(cases),
                    'open_cases': len([c for c in cases if c.status == 'OPEN']),
                    'high_priority': len([c for c in cases if c.priority == 'HIGH']),
                }
            
            # Render HTML template
            html_string = render_to_string(template_name, context)
            
            # Generate PDF using WeasyPrint
            buffer = io.BytesIO()
            HTML(string=html_string, base_url=settings.BASE_DIR).write_pdf(buffer)
            
            buffer.seek(0)
            content = buffer.read()
            
            logger.info(f"Generated PDF report {report.id} using {template_name} (WeasyPrint)")
            return content
            
        except Exception as e:
            logger.error(f"Error generating PDF report {report.id}: {str(e)}", exc_info=True)
            raise
